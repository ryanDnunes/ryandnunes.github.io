from flask import Flask, request, jsonify
import openpyxl
import requests

app = Flask(__name__)

# Carregar a planilha
workbook = openpyxl.load_workbook('numeros.xlsx')
sheet = workbook.active

def get_available_numbers():
    available_numbers = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == True:
            available_numbers.append(row[1])
    return available_numbers

@app.route('/available_numbers', methods=['GET'])
def available_numbers():
    numbers = get_available_numbers()
    return jsonify(numbers)

@app.route('/register', methods=['POST'])
def register():
    data = request.json
    name = data['name']
    email = data.get('email', '')
    whatsapp = data['whatsapp']
    number = int(data['number'])

    for row in sheet.iter_rows(min_row=2):
        if row[1].value == number and row[0].value == True:
            row[0].value = False
            row[2].value = name
            row[3].value = email
            row[4].value = whatsapp
            workbook.save('numeros.xlsx')

            # Enviar mensagem via WhatsApp
            whatsapp_message = f"Nome: {name}\nEmail: {email}\nWhatsApp: {whatsapp}\nNúmero: {number}"
            send_whatsapp_message(whatsapp_message)
            
            return jsonify({'success': True, 'message': 'Registro concluído com sucesso!'})

    return jsonify({'success': False, 'message': 'Número indisponível.'})

def send_whatsapp_message(message):
    url = "https://api.whatsapp.com/send"
    phone_number = "5551985109343"
    payload = {
        'phone': phone_number,
        'text': message
    }
    headers = {
        'Content-Type': 'application/json'
    }
    response = requests.post(url, json=payload, headers=headers)
    return response.status_code == 200

if __name__ == '__main__':
    app.run(debug=True)
